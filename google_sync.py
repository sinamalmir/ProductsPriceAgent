from pathlib import Path
from typing import List

import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

# Path to your service account JSON
SERVICE_ACCOUNT_FILE = Path(
    r"D:\Projects\pdfConvertor\ProductsPriceAgent\iphone-prices.json"
)

# Your target Google Sheet ID
SPREADSHEET_ID = "1q5iauCxDZGcngIzSOqUJ0WbmDUj4pDiiz_CBkW032ds"

# Name of the worksheet/tab inside the spreadsheet
WORKSHEET_NAME = "Sheet1"  # change if needed


def get_gspread_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(
        str(SERVICE_ACCOUNT_FILE),
        scopes=scopes,
    )
    client = gspread.authorize(creds)
    return client


def read_excel_as_2d_list(excel_path: Path) -> List[List[str]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    values: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        row_values = []
        for cell in row:
            if cell is None:
                row_values.append("")
            else:
                row_values.append(str(cell))
        values.append(row_values)

    wb.close()
    return values


def update_google_sheet_from_excel(excel_path: Path):
    """
    Replace the content of the target Google Sheet worksheet
    with the content of the given Excel file.
    """
    client = get_gspread_client()
    sh = client.open_by_key(SPREADSHEET_ID)

    try:
        ws = sh.worksheet(WORKSHEET_NAME)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=WORKSHEET_NAME, rows="1000", cols="30")

    data = read_excel_as_2d_list(excel_path)

    # Clear existing data
    ws.clear()

    if not data:
        return

    # Update starting from A1
    ws.update("A1", data)
