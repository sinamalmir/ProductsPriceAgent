
from pathlib import Path
import re

import win32com.client as win32
from openpyxl import load_workbook


# ------------- SETTINGS (ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø±Ø§ Ø§ÛŒÙ†Ø¬Ø§ Ù…Ø·Ø§Ø¨Ù‚ Ø³ÛŒØ³ØªÙ… Ø®ÙˆØ¯Øª Ø¹ÙˆØ¶ Ú©Ù†) -------------

TEMPLATE_IMPORT_XLSX = Path(r"C:\Users\Administrator\Desktop\pdfConvertor\ProductsPriceAgent\EXCEL\template_pdf_imports.xlsx")
PDF_FOLDER           = Path(r"C:\Users\Administrator\Desktop\pdfConvertor\ProductsPriceAgent\PDF")
OUTPUT_FOLDER        = Path(r"C:\Users\Administrator\Desktop\pdfConvertor\ProductsPriceAgent\converted_excels")

# Ù†Ø§Ù… Query Ø¯Ø± Excel
QUERY_NAME = "Query1"

# ÙØ§ÛŒÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù†Ù‡Ø§ÛŒÛŒ (Ø¬Ø¯ÙˆÙ„ Ø¨Ø²Ø±Ú¯ Ù…Ø¯Ù„â€ŒÙ‡Ø§)
FINAL_TEMPLATE_PATH = Path(r"C:\Users\Administrator\Desktop\pdfConvertor\ProductsPriceAgent\excel_data\FinalTemplate.xlsx")

# Ø³ØªÙˆÙ†ÛŒ Ú©Ù‡ Ù‚ÛŒÙ…Øª Ø¨Ø§ØªØ±ÛŒ Ø±ÙˆÚ©Ø§Ø±ÛŒ Ø¯Ø± Ø¢Ù† Ù†ÙˆØ´ØªÙ‡ Ù…ÛŒâ€ŒØ´ÙˆØ¯
TARGET_COLUMN_HEADER = "Ø¨Ø§Ø·Ø±ÛŒ Ø±ÙˆÚ©Ø§Ø±ÛŒ"
TAG_COLUMN_HEADER    = "ØªÚ¯ Ø¨Ø§Ø·Ø±ÛŒ"              # Ø§Ø² JC PRODUCTS
CAM_FRONT_HEADER     = "Ø¯ÙˆØ±Ø¨ÛŒÙ† Ø¬Ù„Ùˆ Ù¾Ú© Ú©Ø§Ù…Ù„"    # Ø§Ø² apple parts NORMAL (1)
SPEACKERS_HEADER      = "Ø§Ø³Ù¾ÛŒÚ©Ø± Ø¨Ø§Ù„Ø§"         # Ø§Ø² apple parts NORMAL (1)
DOWN_SPEACKERS_HEADER = "Ø§Ø³Ù¾ÛŒÚ©Ø± Ù¾Ø§ÛŒÛŒÙ†"      # Ø§Ø² apple parts NORMAL (1)
FLAT_POWER_HEADER     = "ÙÙ„Øª Ù¾Ø§ÙˆØ±"           # Ø§Ø² apple parts NORMAL (1)
FLAT_VOLUME_HEADER    = "ÙÙ„Øª ÙˆÙ„ÙˆÙ…"           # Ø§Ø² apple parts NORMAL (1)    
FLAT_FLASH_CAMERA_HEADER = "ÙÙ„Øª ÙÙ„Ø´ Ø¯ÙˆØ±Ø¨ÛŒÙ†"     # Ø§Ø² apple parts NORMAL (1)
VIBRATION_HEADER      = "Ù…ÙˆØªÙˆØ± ÙˆÛŒØ¨Ø±Ù‡"           # Ø§Ø² apple parts NORMAL (1)
FRAME_HEADER          = "Ø´Ø§Ø³ÛŒ Ø¨Ø§ ÙÙ„Øª"                # Ø§Ø² apple parts NORMAL (1)
FPC_FLAT_HEADER       = "ÙÙ„Øª Ø§Ù Ù¾ÛŒ Ø³ÛŒ"                # Ø§Ø² apple parts NORMAL (1)
FPC_RECEIVER_HEADER   = "ÙÙ„Øª Ø§Ù Ù¾ÛŒ Ø³ÛŒ Ø¬ÛŒ Ø³ÛŒ"                # Ø§Ø² apple parts NORMAL (1)
FLAT_ANTENNA_HEADER    = "ÙÙ„Øª Ø¢Ù†ØªÙ†"                # Ø§Ø² apple parts NORMAL (1)
FLAT_POWER_WIRELESS_HEADER = "ÙÙ„Øª ÙˆØ§ÛŒØ±Ù„Ø³ Ø´Ø§Ø±Ú˜"    # Ø§Ø² apple parts NORMAL (1)

LENZ_GLASS_HEADER       = "Ø´ÛŒØ´Ù‡ Ù„Ù†Ø² Ø¨Ø§Ø²Ø§Ø±ÛŒ"    # Ø§Ø² apple parts NORMAL (1)
FACE_ID_TAG_HEADER     = "ØªÚ¯ ÙÛŒØ³ Ø§ÛŒØ¯ÛŒ"    # Ø§Ø² apple parts NORMAL (1)
FIX_FACE_ID_HEADER     = "ØªØ¹Ù…ÛŒØ± ÙÛŒØ³ Ø¢ÛŒØ¯ÛŒ"    # Ø§Ø² apple parts NORMAL (1)
FIX_CAMERA_ERROR_HEADER  = "Ø±ÙØ¹ Ø§Ø±ÙˆØ± Ø¯ÙˆØ±Ø¨ÛŒÙ†"    # Ø§Ø² apple parts NORMAL (1)
SHIELD_HEADER          = "Ø´ÛŒÙ„Ø¯"    # Ø§Ø² apple parts NORMAL (1)
ICLOUD_MOTHERBOARD_HEADER = "Ù…Ø§Ø¯Ø±Ø¨Ø±Ø¯ Ø¢ÛŒÚ©Ù„ÙˆØ¯ Ú©Ø§Ù…Ù„"    # Ø§Ø² apple parts NORMAL (1)


# --- Ø¬Ø¯ÛŒØ¯: ÙØ§ÛŒÙ„ JC PRODUCTS Ùˆ Ø³ØªÙˆÙ† Â«ØªÚ¯ Ø¨Ø§Ø·Ø±ÛŒÂ» ---
JC_PRODUCTS_PATH   = Path(r"C:\Users\Administrator\Desktop\pdfConvertor\ProductsPriceAgent\converted_excels\JC PRODUCTS NORMAL.xlsx")
APPLE_PARTS_NORMAL_PATH  = Path(r"C:\Users\Administrator\Desktop\pdfConvertor\ProductsPriceAgent\converted_excels\apple parts NORMAL (1).xlsx")
APPLE_PARTS_RAYAN_PATH  = Path(r"C:\Users\Administrator\Desktop\pdfConvertor\ProductsPriceAgent\converted_excels\Apple_Parts_rayan.xlsx")

# ---------------------------------------------------------------------


def update_query_pdf_path(query, new_pdf_path: Path):
    """Ù…Ø³ÛŒØ± ÙØ§ÛŒÙ„ PDF Ø±Ø§ Ø¯Ø± ÙØ±Ù…ÙˆÙ„ Power Query Ø¹ÙˆØ¶ Ù…ÛŒâ€ŒÚ©Ù†Ø¯."""
    formula = query.Formula
    pdf_str = new_pdf_path.resolve().as_posix()

    new_formula = re.sub(
        r'File\.Contents\(".*?"\)',
        f'File.Contents("{pdf_str}")',
        formula,
    )
    query.Formula = new_formula


def normalize_name(s: str) -> str:
    """Ù†Ø±Ù…Ø§Ù„â€ŒØ³Ø§Ø²ÛŒ Ø§Ø³Ù…â€ŒÙ‡Ø§ Ø¨Ø±Ø§ÛŒ Ù…Ù‚Ø§ÛŒØ³Ù‡ Ø¨Ø¯ÙˆÙ† Ø­Ø³Ø§Ø³ÛŒØª Ø¨Ù‡ ÙØ§ØµÙ„Ù‡/Ø­Ø±ÙˆÙ Ø¨Ø²Ø±Ú¯"""
    return re.sub(r"\s+", " ", s.strip().lower())


MODEL_MAPPING = {
    "8/SE 2020/2022": ["iPhone 8", "iPhone SE (1st Gen)", "iPhone SE (3rd Gen)"],
    "XR": ["iPhone XR"],
    "XS": ["iPhone XS"],
    "XS MAX": ["iPhone XS Max"],
    "11": ["iPhone 11"],
    "11 PRO": ["iPhone 11 Pro"],
    "11 PRO MAX": ["iPhone 11 Pro Max"],
    "12/12 PRO": ["iPhone 12", "iPhone 12 Pro"],
    "12 MINI": ["iPhone 12 mini"],
    "12 PRO MAX": ["iPhone 12 Pro Max"],
    "13": ["iPhone 13"],
    "13 MINI": ["iPhone 13 mini"],
    "13 PRO": ["iPhone 13 Pro"],
    "13 PRO MAX": ["iPhone 13 Pro Max"],
    "14": ["iPhone 14"],
    "14 PLUS": ["iPhone 14 Plus"],
    "14 PRO": ["iPhone 14 Pro"],
    "14 PRO MAX": ["iPhone 14 Pro Max"],
}


# def convert_pdfs_to_excels():
#     """Ù‡Ù…Ù‡ PDFÙ‡Ø§ÛŒ Ù¾ÙˆØ´Ù‡ Ø±Ø§ Ø¨Ø§ Excel Ø¨Ù‡ xlsx ØªØ¨Ø¯ÛŒÙ„ Ù…ÛŒâ€ŒÚ©Ù†Ø¯."""
#     OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

#     excel = win32.Dispatch("Excel.Application")
#     excel.Visible = False

#     try:
#         for pdf_path in PDF_FOLDER.glob("*.pdf"):
#             print(f"Processing PDF -> {pdf_path.name}")

#             wb = excel.Workbooks.Open(str(TEMPLATE_IMPORT_XLSX))
#             query = wb.Queries(QUERY_NAME)

#             update_query_pdf_path(query, pdf_path)

#             wb.RefreshAll()
#             excel.CalculateUntilAsyncQueriesDone()

#             out_path = OUTPUT_FOLDER / f"{pdf_path.stem}.xlsx"
#             wb.SaveAs(str(out_path), FileFormat=51)
#             wb.Close(SaveChanges=False)

#             print(f"Saved Excel -> {out_path}")

#     finally:
#         excel.Quit()

def convert_pdfs_to_excels():
   
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

   
    for f in OUTPUT_FOLDER.glob("*.xlsx"):
        try:
            f.unlink()
            print(f"Deleted old file -> {f}")
        except Exception as e:
            print(f"Could not delete file {f}: {e}")

    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False  # ðŸ”¹ Ø¬Ù„ÙˆÚ¯ÛŒØ±ÛŒ Ø§Ø² Ù‡Ø±Ú¯ÙˆÙ†Ù‡ Ù¾ÛŒØºØ§Ù… Ù…Ø²Ø§Ø­Ù…

    try:
        for pdf_path in PDF_FOLDER.glob("*.pdf"):
            print(f"Processing PDF -> {pdf_path.name}")

            wb = excel.Workbooks.Open(str(TEMPLATE_IMPORT_XLSX))
            query = wb.Queries(QUERY_NAME)

            update_query_pdf_path(query, pdf_path)

            wb.RefreshAll()
            excel.CalculateUntilAsyncQueriesDone()

            out_path = OUTPUT_FOLDER / f"{pdf_path.stem}.xlsx"
            wb.SaveAs(str(out_path), FileFormat=51)
            wb.Close(SaveChanges=False)

            print(f"Saved Excel -> {out_path}")

    finally:
        # Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù†Ø¯Ù† ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø§Ú©Ø³Ù„ Ø¨Ù‡ Ø­Ø§Ù„Øª Ø¹Ø§Ø¯ÛŒ Ùˆ Ø¨Ø³ØªÙ† Ø¨Ø±Ù†Ø§Ù…Ù‡
        try:
            excel.DisplayAlerts = True
        except Exception:
            pass
        excel.Quit()


def build_template_model_index(tmpl_ws):
    """Ø¯ÛŒÚ©Ø´Ù†Ø±ÛŒ: Ø§Ø³Ù… Ù…Ø¯Ù„ Ù†Ø±Ù…Ø§Ù„â€ŒØ´Ø¯Ù‡ â†’ Ø´Ù…Ø§Ø±Ù‡ Ø±Ø¯ÛŒÙ Ø¯Ø± ØªÙ…Ù¾Ù„ÛŒØª"""
    index = {}
    for row in range(2, tmpl_ws.max_row + 1):
        val = tmpl_ws.cell(row=row, column=1).value
        if not val:
            continue
        index[normalize_name(str(val))] = row
    return index


def find_column_by_header(ws, header_text: str):
    """Ø´Ù…Ø§Ø±Ù‡ Ø³ØªÙˆÙ† Ø±Ø§ Ø¨Ø± Ø§Ø³Ø§Ø³ Ù…ØªÙ† Ù‡Ø¯Ø± Ø¯Ø± Ø±Ø¯ÛŒÙ Û± Ø¨Ø±Ù…ÛŒâ€ŒÚ¯Ø±Ø¯Ø§Ù†Ø¯."""
    for cell in ws[1]:
        if str(cell.value).strip() == header_text:
            return cell.column
    return None


# --------- 1) Ù¾Ø± Ú©Ø±Ø¯Ù† Ø³ØªÙˆÙ† Â«Ø¨Ø§Ø·Ø±ÛŒ Ø±ÙˆÚ©Ø§Ø±ÛŒÂ» Ø§Ø² cell HIGH CAPACITY ----------

def fill_template_from_converted_excel(converted_xlsx: Path):
    print(f"Reading data from: {converted_xlsx}")

    data_wb = load_workbook(converted_xlsx, data_only=True)
    data_ws = data_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    target_col_idx = find_column_by_header(tmpl_ws, TARGET_COLUMN_HEADER)
    if target_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{TARGET_COLUMN_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in data_ws.iter_rows(
        min_row=6, min_col=3, max_col=4, values_only=True
    ):
        if not model_val or not price_val:
            continue

        if str(price_val).strip() == "-":
            continue

        key = str(model_val).strip().upper()
        mapped_models = MODEL_MAPPING.get(key)

        if not mapped_models:
            generic_name = "iPhone " + str(model_val).strip()
            norm = normalize_name(generic_name)
            if norm in template_model_rows:
                mapped_models = [generic_name]
            else:
                print(f"Ù…Ø¯Ù„ Ù†Ø§Ø´Ù†Ø§Ø®ØªÙ‡ Ø¯Ø± Ø¯ÛŒØªØ§ (cell HIGH CAPACITY): {model_val}  (Ø±Ø¯ Ø´Ø¯)")
                continue

        for tmpl_model_name in mapped_models:
            row_idx = template_model_rows.get(normalize_name(tmpl_model_name))
            if not row_idx:
                print(f"Ù…Ø¯Ù„ '{tmpl_model_name}' Ø¯Ø± ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯ (Ø§Ø² Ø±ÙˆÛŒ {model_val})")
                continue

            tmpl_ws.cell(row=row_idx, column=target_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with prices from {converted_xlsx.name}")


# --------- 2) Ù¾Ø± Ú©Ø±Ø¯Ù† Ø³ØªÙˆÙ† Â«ØªÚ¯ Ø¨Ø§Ø·Ø±ÛŒÂ» Ø§Ø² JC PRODUCTS ----------

def fill_template_from_jc_products(jc_xlsx: Path):
    print(f"Reading JC Products from: {jc_xlsx}")

    jc_wb = load_workbook(jc_xlsx, data_only=True)
    jc_ws = jc_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    tag_col_idx = find_column_by_header(tmpl_ws, TAG_COLUMN_HEADER)
    if tag_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{TAG_COLUMN_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # E7:F28
    for model_val, tag_val in jc_ws.iter_rows(
        min_row=7, max_row=28, min_col=5, max_col=6, values_only=True
    ):
        if not model_val or not tag_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"Ù…Ø¯Ù„ '{model_val}' Ø¯Ø± ØªÙ…Ù¾Ù„ÛŒØª Ø¨Ø±Ø§ÛŒ ØªÚ¯ Ø¨Ø§ØªØ±ÛŒ Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯ (JC PRODUCTS)")
            continue

        tmpl_ws.cell(row=row_idx, column=tag_col_idx).value = tag_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with tags from {jc_xlsx.name}")


# --------- 3) Ù¾Ø± Ú©Ø±Ø¯Ù† Ø³ØªÙˆÙ† Â«Ø¯ÙˆØ±Ø¨ÛŒÙ† Ø¬Ù„Ùˆ Ù¾Ú© Ú©Ø§Ù…Ù„Â» Ø§Ø² apple parts NORMAL (1) ----------

def fill_template_from_apple_parts_normal(apple_xlsx: Path):
    """
    Ø§Ø² ÙØ§ÛŒÙ„ apple parts NORMAL (1).xlsx Ø³Ù„ÙˆÙ„â€ŒÙ‡Ø§ÛŒ C4:D43 Ø±Ø§ Ù…ÛŒâ€ŒØ®ÙˆØ§Ù†Ø¯
    Ùˆ Ù…Ù‚Ø¯Ø§Ø± Ø³ØªÙˆÙ† D Ø±Ø§ Ø¯Ø± Ø³ØªÙˆÙ† Â«Ø¯ÙˆØ±Ø¨ÛŒÙ† Ø¬Ù„Ùˆ Ù¾Ú© Ú©Ø§Ù…Ù„Â» ØªÙ…Ù¾Ù„ÛŒØª Ù…ÛŒâ€ŒÙ†ÙˆÛŒØ³Ø¯.
    """
    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, CAM_FRONT_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{CAM_FRONT_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # C4:D43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=3, max_col=4, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"Ù…Ø¯Ù„ '{model_val}' Ø¯Ø± ØªÙ…Ù¾Ù„ÛŒØª Ø¨Ø±Ø§ÛŒ Ø¯ÙˆØ±Ø¨ÛŒÙ† Ø¬Ù„Ùˆ Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯ (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


def fill_template_from_apple_parts_normal_downSpeackers(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, SPEACKERS_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{SPEACKERS_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # S4:T43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=19, max_col=20, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"Ù…Ø¯Ù„ '{model_val}' Ø¯Ø± ØªÙ…Ù¾Ù„ÛŒØª Ø¨Ø±Ø§ÛŒ Ø¯ÙˆØ±Ø¨ÛŒÙ† Ø¬Ù„Ùˆ Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯ (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


def fill_template_from_apple_parts_normal_speakers(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, DOWN_SPEACKERS_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{DOWN_SPEACKERS_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # Q4:R43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=17, max_col=18, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"Ù…Ø¯Ù„ '{model_val}' Ø¯Ø± ØªÙ…Ù¾Ù„ÛŒØª Ø¨Ø±Ø§ÛŒ Ø¯ÙˆØ±Ø¨ÛŒÙ† Ø¬Ù„Ùˆ Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯ (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


def fill_template_from_apple_parts_normal_flat_power(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, FLAT_POWER_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{FLAT_POWER_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # G4:H43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=7, max_col=8, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")

def fill_template_from_apple_parts_normal_flat_power_volume(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, FLAT_VOLUME_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{FLAT_VOLUME_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # W4:X43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=23, max_col=24, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")




def fill_template_from_apple_parts_normal_flat_flash_camera(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, FLAT_FLASH_CAMERA_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{FLAT_FLASH_CAMERA_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # AE4:AF43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=31, max_col=32, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")



def fill_template_from_apple_parts_normal_vibration(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, VIBRATION_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{VIBRATION_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # U4:V43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=21, max_col=22, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


#rayan pdf file
def fill_template_from_apple_parts_rayan_frame(apple_xlsx: Path):

    print(f"Reading Apple Parts rayan from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, FRAME_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{FRAME_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # M4:N43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=38, min_col=13, max_col=14, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


def fill_template_from_apple_parts_normal_fpc_flat(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, VIBRATION_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{VIBRATION_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # AM4:AN3
    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=39, max_col=40, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


def fill_template_from_apple_parts_normal_fpc_receiver(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, FPC_RECEIVER_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{FPC_RECEIVER_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # G4:HN3
    for model_val, price_val in ap_ws.iter_rows(
        min_row=7, max_row=28, min_col=7, max_col=8, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


def fill_template_from_apple_parts_normal_flat_antenna(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, FLAT_ANTENNA_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{FLAT_ANTENNA_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # AW4:AX28
    for model_val, price_val in ap_ws.iter_rows(
        min_row=5, max_row=43, min_col=49, max_col=50, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")



def fill_template_from_apple_parts_normal_lenz_glass(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, FLAT_POWER_WIRELESS_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{FLAT_POWER_WIRELESS_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # K6:L43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=6, max_row=43, min_col=11, max_col=12, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


def fill_template_from_apple_parts_normal_flat_wireless(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, LENZ_GLASS_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{LENZ_GLASS_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # O6:P43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=6, max_row=43, min_col=15, max_col=16, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


def fill_template_from_apple_parts_normal_face_id_tag(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, FACE_ID_TAG_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{FACE_ID_TAG_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # Q7:R28
    for model_val, price_val in ap_ws.iter_rows(
        min_row=7, max_row=28, min_col=17, max_col=18, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")



def fill_template_from_apple_parts_normal_fix_face_id(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, FIX_FACE_ID_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{FIX_FACE_ID_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # C7:D28
    for model_val, price_val in ap_ws.iter_rows(
        min_row=7, max_row=28, min_col=3, max_col=4, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


def fill_template_from_apple_parts_normal_fix_camera_error(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, FIX_CAMERA_ERROR_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{FIX_CAMERA_ERROR_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # M7:N28
    for model_val, price_val in ap_ws.iter_rows(
        min_row=7, max_row=28, min_col=13, max_col=14, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")


def fill_template_from_apple_parts_normal_shield(apple_xlsx: Path):

    print(f"Reading Apple Parts Normal from: {apple_xlsx}")

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    cam_col_idx = find_column_by_header(tmpl_ws, SHIELD_HEADER)
    if cam_col_idx is None:
        raise ValueError(f"Ø³ØªÙˆÙ†ÛŒ Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† '{SHIELD_HEADER}' Ø¯Ø± Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„ ØªÙ…Ù¾Ù„ÛŒØª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

    template_model_rows = build_template_model_index(tmpl_ws)

    # Y6:Z43
    for model_val, price_val in ap_ws.iter_rows(
        min_row=6, max_row=43, min_col=25, max_col=26, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            print(f"This model : '{model_val}' Not found in file (apple parts NORMAL)")
            continue

        tmpl_ws.cell(row=row_idx, column=cam_col_idx).value = price_val

    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    print(f"Template updated with front camera prices from {apple_xlsx.name}")




def main():
   # if u need to convert pdfs to excel every time --> uncommnet this code 

    convert_pdfs_to_excels()

   
    # converted_file = OUTPUT_FOLDER / "cell  HIGH CAPACITY.xlsx"  
    # fill_template_from_converted_excel(converted_file)

   
    # fill_template_from_jc_products(JC_PRODUCTS_PATH)

  
    # fill_template_from_apple_parts_normal(APPLE_PARTS_NORMAL_PATH)

    # fill_template_from_apple_parts_normal_speakers(APPLE_PARTS_NORMAL_PATH)

    # fill_template_from_apple_parts_normal_downSpeackers(APPLE_PARTS_NORMAL_PATH)
    # fill_template_from_apple_parts_normal_flat_power_volume(APPLE_PARTS_NORMAL_PATH)
    # fill_template_from_apple_parts_normal_flat_flash_camera(APPLE_PARTS_NORMAL_PATH)
    # fill_template_from_apple_parts_normal_vibration(APPLE_PARTS_NORMAL_PATH)

    # fill_template_from_apple_parts_rayan_frame(APPLE_PARTS_RAYAN_PATH)

    # fill_template_from_apple_parts_normal_fpc_flat(APPLE_PARTS_NORMAL_PATH)

    # fill_template_from_apple_parts_normal_fpc_receiver(JC_PRODUCTS_PATH)
    # fill_template_from_apple_parts_normal_flat_antenna(APPLE_PARTS_NORMAL_PATH)

    # fill_template_from_apple_parts_normal_flat_wireless(APPLE_PARTS_NORMAL_PATH)

    # fill_template_from_apple_parts_normal_flat_wireless(APPLE_PARTS_NORMAL_PATH)

    # fill_template_from_apple_parts_normal_face_id_tag(JC_PRODUCTS_PATH)

    # fill_template_from_apple_parts_normal_fix_face_id(JC_PRODUCTS_PATH)

    # fill_template_from_apple_parts_normal_fix_camera_error(JC_PRODUCTS_PATH)

    # fill_template_from_apple_parts_normal_shield(APPLE_PARTS_NORMAL_PATH)

if __name__ == "__main__":
    main()