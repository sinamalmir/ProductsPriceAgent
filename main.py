from pathlib import Path
import re
import logging
import asyncio
from typing import Dict, List

import win32com.client as win32
from openpyxl import load_workbook

from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

# ====================== TELEGRAM SETTINGS ======================

BOT_TOKEN = "7266866129:AAHkXCVpqoy4uFLZzHlSTS-ycI_C-1FKKmQ"

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ======================== FILE SETTINGS ========================

TEMPLATE_IMPORT_XLSX = Path(
    r"D:\Projects\pdfConvertor\ProductsPriceAgent\EXCEL\template_pdf_imports.xlsx"
)
PDF_FOLDER = Path(
    r"D:\Projects\pdfConvertor\ProductsPriceAgent\PDF"
)
OUTPUT_FOLDER = Path(
    r"D:\Projects\pdfConvertor\ProductsPriceAgent\converted_excels"
)

QUERY_NAME = "Query1"

FINAL_TEMPLATE_PATH = Path(
    r"D:\Projects\pdfConvertor\ProductsPriceAgent\excel_data\FinalTemplate.xlsx"
)

# Column headers in final template
TARGET_COLUMN_HEADER = "باطری روکاری"
TAG_COLUMN_HEADER = "تگ باطری"
CAM_FRONT_HEADER = "دوربین جلو پک کامل"
SPEACKERS_HEADER = "اسپیکر بالا"
DOWN_SPEACKERS_HEADER = "اسپیکر پایین"
FLAT_POWER_HEADER = "فلت پاور"
FLAT_VOLUME_HEADER = "فلت ولوم"
FLAT_FLASH_CAMERA_HEADER = "فلت فلش دوربین"
VIBRATION_HEADER = "موتور ویبره"
FRAME_HEADER = "شاسی با فلت"
FPC_FLAT_HEADER = "فلت اف پی سی"
FPC_RECEIVER_HEADER = "فلت اف پی سی جی سی"
FLAT_ANTENNA_HEADER = "فلت آنتن"
FLAT_POWER_WIRELESS_HEADER = "فلت وایرلس شارژ"

LENZ_GLASS_HEADER = "شیشه لنز بازاری"
FACE_ID_TAG_HEADER = "تگ فیس ایدی"
FIX_FACE_ID_HEADER = "تعمیر فیس آیدی"
FIX_CAMERA_ERROR_HEADER = "رفع ارور دوربین"
SHIELD_HEADER = "شیلد"
ICLOUD_MOTHERBOARD_HEADER = "مادربرد آیکلود کامل"

LCD_CHANGED_GLASS_HEADER = "Lcd چنج گلس یا فلت خورده"
LCD_COPY_HEADER = "Lcd copy"
LCD_1_CHANGED_FRAME_HEADER = "Lcd1.1 فریم تعویض"
CHARGE_FLAT_HEADER = "فلت شارژ"
LCD_ROKARI_WITH_POLISH_HEADER = "Lcd روکاری بازاری,پولیش دارد"
LCD_ROKARI_WITHOUT_POLISH_HEADER = "Lcd روکاری بدون پولیش"
LCD_NEW_APPLE_HEADER = "Lcd new apple"
LCD_USED_NORMAL_HEADER = "Lcd used normal"
BACK_CAMERA_HEADER = "دوربین عقب"

# ========================== BUFFERS ============================

USER_PDF_BUFFER: Dict[int, List[Path]] = {}
USER_TIMER: Dict[int, asyncio.Task] = {}

# ======================== CORE HELPERS =========================


def update_query_pdf_path(query, new_pdf_path: Path):
    formula = query.Formula
    pdf_str = new_pdf_path.resolve().as_posix()

    new_formula = re.sub(
        r'File\.Contents\(".*?"\)',
        f'File.Contents("{pdf_str}")',
        formula,
    )
    query.Formula = new_formula


def normalize_name(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


MODEL_MAPPING = {
    "8/SE 2020/2022": ["iPhone 8", "iPhone SE (1st Gen)", "iPhone SE (3rd Gen)"],
    "XR": ["iPhone XR"],
    "XS": ["iPhone XS"],
    "XS MAX": ["iPhone XS Max"],
    "11": ["iPhone 11"],
    "11 PRO": ["iPhone 11 Pro"],
    "11 PRO MAX": ["iPhone 11 Pro Max"],
    "12/12 PRO": ["iPhone 12", "iPhone 12 Pro"],
    "12 MINI": ["iPhone 12 mini"],
    "12 PRO MAX": ["iPhone 12 Pro Max"],
    "13": ["iPhone 13"],
    "13 MINI": ["iPhone 13 mini"],
    "13 PRO": ["iPhone 13 Pro"],
    "13 PRO MAX": ["iPhone 13 Pro Max"],
    "14": ["iPhone 14"],
    "14 PLUS": ["iPhone 14 Plus"],
    "14 PRO": ["iPhone 14 Pro"],
    "14 PRO MAX": ["iPhone 14 Pro Max"],
}


def build_template_model_index(tmpl_ws):
    index = {}
    for row in range(2, tmpl_ws.max_row + 1):
        val = tmpl_ws.cell(row=row, column=1).value
        if not val:
            continue
        index[normalize_name(str(val))] = row
    return index


def find_column_by_header(ws, header_text: str):
    for cell in ws[1]:
        if str(cell.value).strip() == header_text:
            return cell.column
    return None


def append_price(existing_value, new_price):
    new_str = str(new_price).strip()
    if new_str == "":
        return existing_value

    if existing_value is None or str(existing_value).strip() == "":
        return new_str

    parts = [p.strip() for p in str(existing_value).split("/") if p.strip()]
    if new_str in parts:
        return " / ".join(parts)

    parts.append(new_str)
    return " / ".join(parts)


# ================= PDF -> EXCEL (ONE FILE) =====================


def convert_pdf_to_excel(pdf_path: Path) -> Path:
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

    if not TEMPLATE_IMPORT_XLSX.exists():
        raise FileNotFoundError(
            f"Template import XLSX not found: {TEMPLATE_IMPORT_XLSX}"
        )

    out_path = OUTPUT_FOLDER / f"{pdf_path.stem}.xlsx"

    if out_path.exists():
        try:
            out_path.unlink()
            logger.info("Deleted old converted file: %s", out_path)
        except Exception as e:
            logger.warning("Could not delete %s: %s", out_path, e)

    logger.info("Converting PDF to Excel: %s -> %s", pdf_path, out_path)

    excel = win32.Dispatch("Excel.Application")

    try:
        wb = excel.Workbooks.Open(str(TEMPLATE_IMPORT_XLSX))
        query = wb.Queries(QUERY_NAME)

        update_query_pdf_path(query, pdf_path)

        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()

        wb.SaveAs(str(out_path), FileFormat=51)
        wb.Close(SaveChanges=False)

        logger.info("Saved Excel: %s", out_path)
    finally:
        excel.Quit()

    return out_path


# ========================= FILL FUNCTIONS =========================

def fill_template_from_converted_excel(converted_xlsx: Path):
    logger.info("Reading battery prices from: %s", converted_xlsx)

    data_wb = load_workbook(converted_xlsx, data_only=True)
    data_ws = data_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    target_col_idx = find_column_by_header(tmpl_ws, TARGET_COLUMN_HEADER)
    if target_col_idx is None:
        raise ValueError(
            f"Header '{TARGET_COLUMN_HEADER}' not found in final template."
        )

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in data_ws.iter_rows(
        min_row=6, min_col=3, max_col=4, values_only=True
    ):
        if not model_val or not price_val:
            continue

        if str(price_val).strip() == "-":
            continue

        key = str(model_val).strip().upper()
        mapped_models = MODEL_MAPPING.get(key)

        if not mapped_models:
            generic_name = "iPhone " + str(model_val).strip()
            norm = normalize_name(generic_name)
            if norm in template_model_rows:
                mapped_models = [generic_name]
            else:
                logger.warning("Unknown model in battery file: %s", model_val)
                continue

        for tmpl_model_name in mapped_models:
            row_idx = template_model_rows.get(normalize_name(tmpl_model_name))
            if not row_idx:
                logger.warning(
                    "Model '%s' not found in final template (from %s)",
                    tmpl_model_name,
                    model_val,
                )
                continue

            tmpl_ws.cell(row=row_idx, column=target_col_idx).value = price_val

    data_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Battery prices updated from %s", converted_xlsx.name)


def fill_template_from_jc_products(jc_xlsx: Path):
    logger.info("Reading JC Products from: %s", jc_xlsx)

    jc_wb = load_workbook(jc_xlsx, data_only=True)
    jc_ws = jc_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    tag_col_idx = find_column_by_header(tmpl_ws, TAG_COLUMN_HEADER)
    if tag_col_idx is None:
        raise ValueError(f"Header '{TAG_COLUMN_HEADER}' not found in final template.")

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, tag_val in jc_ws.iter_rows(
        min_row=7, max_row=28, min_col=5, max_col=6, values_only=True
    ):
        if not model_val or not tag_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for battery tag not found in final template", model_val
            )
            continue

        tmpl_ws.cell(row=row_idx, column=tag_col_idx).value = tag_val

    jc_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Battery tags updated from %s", jc_xlsx.name)


def fill_template_from_apple_parts_normal(apple_xlsx: Path):
    logger.info("Reading Apple Parts Normal (front camera) from: %s", apple_xlsx)

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    col_idx = find_column_by_header(tmpl_ws, CAM_FRONT_HEADER)
    if col_idx is None:
        raise ValueError(f"Header '{CAM_FRONT_HEADER}' not found in final template.")

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=3, max_col=4, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for front camera not found in final template", model_val
            )
            continue

        tmpl_ws.cell(row=row_idx, column=col_idx).value = price_val

    ap_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Front camera prices updated from %s", apple_xlsx.name)


def fill_template_from_apple_parts_normal_downSpeackers(apple_xlsx: Path):
    logger.info("Reading Apple Parts Normal (upper speaker) from: %s", apple_xlsx)

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    col_idx = find_column_by_header(tmpl_ws, SPEACKERS_HEADER)
    if col_idx is None:
        raise ValueError(f"Header '{SPEACKERS_HEADER}' not found in final template.")

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=19, max_col=20, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for upper speaker not found in final template", model_val
            )
            continue

        tmpl_ws.cell(row=row_idx, column=col_idx).value = price_val

    ap_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Upper speaker prices updated from %s", apple_xlsx.name)


def fill_template_from_apple_parts_normal_speakers(apple_xlsx: Path):
    logger.info("Reading Apple Parts Normal (lower speaker) from: %s", apple_xlsx)

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    col_idx = find_column_by_header(tmpl_ws, DOWN_SPEACKERS_HEADER)
    if col_idx is None:
        raise ValueError(
            f"Header '{DOWN_SPEACKERS_HEADER}' not found in final template."
        )

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=17, max_col=18, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for lower speaker not found in final template", model_val
            )
            continue

        tmpl_ws.cell(row=row_idx, column=col_idx).value = price_val

    ap_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Lower speaker prices updated from %s", apple_xlsx.name)


def fill_template_from_apple_parts_normal_flat_power(apple_xlsx: Path):
    logger.info("Reading Apple Parts Normal (flat power) from: %s", apple_xlsx)

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    col_idx = find_column_by_header(tmpl_ws, FLAT_POWER_HEADER)
    if col_idx is None:
        raise ValueError(f"Header '{FLAT_POWER_HEADER}' not found in final template.")

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=7, max_col=8, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for flat power not found in final template", model_val
            )
            continue

        tmpl_ws.cell(row=row_idx, column=col_idx).value = price_val

    ap_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Flat power prices updated from %s", apple_xlsx.name)


def fill_template_from_apple_parts_normal_flat_power_volume(apple_xlsx: Path):
    logger.info("Reading Apple Parts Normal (flat volume) from: %s", apple_xlsx)

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    col_idx = find_column_by_header(tmpl_ws, FLAT_VOLUME_HEADER)
    if col_idx is None:
        raise ValueError(
            f"Header '{FLAT_VOLUME_HEADER}' not found in final template."
        )

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=23, max_col=24, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for flat volume not found in final template", model_val
            )
            continue

        tmpl_ws.cell(row=row_idx, column=col_idx).value = price_val

    ap_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Flat volume prices updated from %s", apple_xlsx.name)


def fill_template_from_apple_parts_normal_flat_flash_camera(apple_xlsx: Path):
    logger.info("Reading Apple Parts Normal (flash flat) from: %s", apple_xlsx)

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    col_idx = find_column_by_header(tmpl_ws, FLAT_FLASH_CAMERA_HEADER)
    if col_idx is None:
        raise ValueError(
            f"Header '{FLAT_FLASH_CAMERA_HEADER}' not found in final template."
        )

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=31, max_col=32, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for flash flat not found in final template", model_val
            )
            continue

        tmpl_ws.cell(row=row_idx, column=col_idx).value = price_val

    ap_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Flash flat prices updated from %s", apple_xlsx.name)


def fill_template_from_apple_parts_normal_vibration(apple_xlsx: Path):
    logger.info("Reading Apple Parts Normal (vibration) from: %s", apple_xlsx)

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    col_idx = find_column_by_header(tmpl_ws, VIBRATION_HEADER)
    if col_idx is None:
        raise ValueError(
            f"Header '{VIBRATION_HEADER}' not found in final template."
        )

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=43, min_col=21, max_col=22, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for vibration not found in final template", model_val
            )
            continue

        tmpl_ws.cell(row=row_idx, column=col_idx).value = price_val

    ap_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Vibration prices updated from %s", apple_xlsx.name)


def fill_template_from_apple_parts_rayan_frame(apple_xlsx: Path):
    logger.info("Reading Apple Parts Rayan (frame) from: %s", apple_xlsx)

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    col_idx = find_column_by_header(tmpl_ws, FRAME_HEADER)
    if col_idx is None:
        raise ValueError(f"Header '{FRAME_HEADER}' not found in final template.")

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=38, min_col=13, max_col=14, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for frame not found in final template (Rayan)", model_val
            )
            continue

        tmpl_ws.cell(row=row_idx, column=col_idx).value = price_val

    ap_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Frame prices updated from %s", apple_xlsx.name)


def fill_template_from_apple_parts_normal_charge_flat(apple_xlsx: Path):
    logger.info("Reading Apple Parts Normal (charge flat) from: %s", apple_xlsx)

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    col_idx = find_column_by_header(tmpl_ws, CHARGE_FLAT_HEADER)
    if col_idx is None:
        raise ValueError(f"Header '{CHARGE_FLAT_HEADER}' not found in final template.")

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in ap_ws.iter_rows(
        min_row=6, max_row=43, min_col=5, max_col=6, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for charge flat not found in final template (Normal)",
                model_val,
            )
            continue

        cell = tmpl_ws.cell(row=row_idx, column=col_idx)
        cell.value = append_price(cell.value, price_val)

    ap_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Charge-flat prices (Normal) updated from %s", apple_xlsx.name)


def fill_template_from_apple_parts_rayan_charge_flat(apple_xlsx: Path):
    logger.info("Reading Apple Parts Rayan (charge flat) from: %s", apple_xlsx)

    ap_wb = load_workbook(apple_xlsx, data_only=True)
    ap_ws = ap_wb.active

    tmpl_wb = load_workbook(FINAL_TEMPLATE_PATH)
    tmpl_ws = tmpl_wb.active

    col_idx = find_column_by_header(tmpl_ws, CHARGE_FLAT_HEADER)
    if col_idx is None:
        raise ValueError(f"Header '{CHARGE_FLAT_HEADER}' not found in final template.")

    template_model_rows = build_template_model_index(tmpl_ws)

    for model_val, price_val in ap_ws.iter_rows(
        min_row=4, max_row=38, min_col=11, max_col=12, values_only=True
    ):
        if not model_val or not price_val:
            continue

        norm_model = normalize_name(str(model_val))
        row_idx = template_model_rows.get(norm_model)

        if not row_idx:
            generic_name = "iPhone " + str(model_val).strip()
            row_idx = template_model_rows.get(normalize_name(generic_name))

        if not row_idx:
            logger.warning(
                "Model '%s' for charge flat not found in final template (Rayan)",
                model_val,
            )
            continue

        cell = tmpl_ws.cell(row=row_idx, column=col_idx)
        cell.value = append_price(cell.value, price_val)

    ap_wb.close()
    tmpl_wb.save(FINAL_TEMPLATE_PATH)
    tmpl_wb.close()
    logger.info("Charge-flat prices (Rayan) updated from %s", apple_xlsx.name)


# ==================== HIGH-LEVEL PIPELINE (BATCH) ====================


def run_full_price_pipeline_for_batch(pdf_paths: List[Path]):
    logger.info("Starting price pipeline for batch: %d file(s)", len(pdf_paths))

    if not FINAL_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Final template XLSX not found: {FINAL_TEMPLATE_PATH}"
        )

    for pdf_path in pdf_paths:
        converted_file = convert_pdf_to_excel(pdf_path)
        s = pdf_path.stem.strip().lower()

        if "cell" in s and "high" in s:
            fill_template_from_converted_excel(converted_file)

        elif "jc" in s and "product" in s:
            fill_template_from_jc_products(converted_file)

        elif "apple" in s and "normal" in s:
            fill_template_from_apple_parts_normal(converted_file)
            fill_template_from_apple_parts_normal_speakers(converted_file)
            fill_template_from_apple_parts_normal_downSpeackers(converted_file)
            fill_template_from_apple_parts_normal_flat_power(converted_file)
            fill_template_from_apple_parts_normal_flat_power_volume(converted_file)
            fill_template_from_apple_parts_normal_flat_flash_camera(converted_file)
            fill_template_from_apple_parts_normal_vibration(converted_file)
            fill_template_from_apple_parts_normal_charge_flat(converted_file)

        elif "rayan" in s:
            fill_template_from_apple_parts_rayan_frame(converted_file)
            fill_template_from_apple_parts_rayan_charge_flat(converted_file)

        else:
            logger.info("No rule defined for PDF name: %s", pdf_path.name)

    logger.info("Batch price pipeline finished.")


# ====================== TELEGRAM BOT PART ======================


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "سلام\n"
        "فایل‌های PDF لیست قیمت را ارسال کنید.\n"
        "می‌توانید چند فایل پشت سر هم بفرستید؛ پس از دریافت همهٔ فایل‌ها، پردازش انجام می‌شود و یک فایل اکسل نهایی برای شما ارسال خواهد شد."
    )


async def process_user_pdfs(user_id: int, message):
    pdf_list = USER_PDF_BUFFER.get(user_id, [])
    if not pdf_list:
        return

    USER_PDF_BUFFER[user_id] = []
    USER_TIMER.pop(user_id, None)

    try:
        await message.reply_text(
            "فایل‌ها دریافت شد. در حال پردازش اطلاعات، لطفاً منتظر بمانید..."
        )

        run_full_price_pipeline_for_batch(pdf_list)

        if FINAL_TEMPLATE_PATH.exists():
            with open(FINAL_TEMPLATE_PATH, "rb") as f:
                await message.reply_document(
                    document=f,
                    filename=FINAL_TEMPLATE_PATH.name,
                    caption="فایل اکسل نهایی آماده است.",
                )
        else:
            await message.reply_text(
                "در حین پردازش، فایل نهایی پیدا نشد. لطفاً تنظیمات سرور بررسی شود."
            )

    except FileNotFoundError as e:
        logger.exception("File not found in pipeline")
        await message.reply_text(
            "در حین پردازش، فایل‌های لازم پیدا نشدند.\n"
            f"جزئیات: {e}"
        )
    except Exception:
        logger.exception("Unexpected error in pipeline")
        await message.reply_text(
            "در حین پردازش فایل‌ها خطایی رخ داد.\n"
            "لطفاً وضعیت فایل‌ها را بررسی کنید یا با پشتیبان سیستم هماهنگ شوید."
        )


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message = update.message
    doc = message.document

    if not doc:
        return

    if doc.mime_type != "application/pdf" and not doc.file_name.lower().endswith(".pdf"):
        await message.reply_text("لطفاً فقط فایل‌های PDF ارسال شود.")
        return

    user_id = message.from_user.id
    PDF_FOLDER.mkdir(parents=True, exist_ok=True)

    filename = doc.file_name or f"{doc.file_unique_id}.pdf"
    pdf_path = PDF_FOLDER / filename

    try:
        tg_file = await doc.get_file()
        await tg_file.download_to_drive(str(pdf_path))
        logger.info("Downloaded PDF for user %s to %s", user_id, pdf_path)
    except Exception:
        logger.exception("Error downloading file")
        await message.reply_text(
            "در دریافت فایل مشکلی پیش آمد. لطفاً دوباره تلاش کنید."
        )
        return

    USER_PDF_BUFFER.setdefault(user_id, []).append(pdf_path)

    old_task = USER_TIMER.get(user_id)
    if old_task is not None and not old_task.done():
        old_task.cancel()

    async def timer_task():
        try:
            await asyncio.sleep(3)
            await process_user_pdfs(user_id, message)
        except asyncio.CancelledError:
            return

    USER_TIMER[user_id] = asyncio.create_task(timer_task())

    await message.reply_text(
        "فایل دریافت شد.\n"
        "در صورت نیاز، فایل‌های دیگر را هم ارسال کنید.\n"
        "اگر فایل جدیدی ارسال نشود، پس از چند ثانیه پردازش به‌صورت خودکار آغاز می‌شود."
    )


def main():
    application = ApplicationBuilder().token(BOT_TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    logger.info("Bot is starting...")
    application.run_polling()


if __name__ == "__main__":
    main()
